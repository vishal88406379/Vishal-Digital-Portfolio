import pandas as pd
import os
from tkinter import Tk, Label, Button, filedialog
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side

def select_input_folder():
    folder_selected = filedialog.askdirectory()
    input_folder_label.config(text=folder_selected)

def select_output_folder():
    folder_selected = filedialog.askdirectory()
    output_folder_label.config(text=folder_selected)

def process_files():
    input_folder = input_folder_label.cget("text")
    output_folder = output_folder_label.cget("text")

    if not input_folder or not output_folder:
        print("Please select both input and output folders.")
        return

    for filename in os.listdir(input_folder):
        if filename.endswith('.xls'):
            file_path = os.path.join(input_folder, filename)
            df = pd.read_excel(file_path)

            # Creating a new Excel writer object
            output_file_path = os.path.join(output_folder, filename.replace('.xls', '.xlsx'))
            writer = pd.ExcelWriter(output_file_path, engine='openpyxl')
            df.to_excel(writer, index=False, sheet_name='Sheet1')

            # Accessing the workbook and the active sheet
            worksheet = writer.sheets['Sheet1']

            # Set header color to yellow and center align
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            for cell in worksheet[1]:  # Assuming the first row is the header
                cell.fill = yellow_fill
                cell.alignment = Alignment(horizontal='center')

            # Center align all cells and apply border
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))

            for row in worksheet.iter_rows(min_row=2):  # Start from the second row
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border

            # AutoFit column width
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]  # Convert to list to access cells
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)  # Add a little extra space
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

            writer.close()  # This will save the file

            # Move file to dealer-specific folder
            for dealer in df['Dealer'].unique():  # Ensure this matches your column name
                dealer_folder = os.path.join(output_folder, dealer)
                os.makedirs(dealer_folder, exist_ok=True)
                dealer_file_path = os.path.join(dealer_folder, filename.replace('.xls', '.xlsx'))
                
                # Move the saved file to the dealer's folder
                os.rename(output_file_path, dealer_file_path)

    print("Processing completed.")

# GUI Setup
root = Tk()
root.title("Excel File Processor")

Label(root, text="Select Input Folder:").pack()
input_folder_label = Label(root, text="", width=50)
input_folder_label.pack()

Button(root, text="Browse", command=select_input_folder).pack()

Label(root, text="Select Output Folder:").pack()
output_folder_label = Label(root, text="", width=50)
output_folder_label.pack()

Button(root, text="Browse", command=select_output_folder).pack()

Button(root, text="Run", command=process_files).pack()

root.mainloop()
