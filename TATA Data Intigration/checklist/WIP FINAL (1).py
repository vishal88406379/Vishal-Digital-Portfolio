import tkinter as tk
from tkinter import filedialog, StringVar
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import calendar
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font

# Global variables to hold the file paths (initialized as empty strings)
LOCATION_MASTER_PATH = ""
PART_MASTER_PATH = ""

# Function to clean the data and add required columns
def clean_data(input_files, output_path):
    global LOCATION_MASTER_PATH, PART_MASTER_PATH
    
    # Load the location master file
    location_master = pd.read_excel(LOCATION_MASTER_PATH)
    location_master.rename(columns={"Code": "Division", "Final Location": "Location"}, inplace=True)

    # Load the part master file
    part_master = pd.read_excel(PART_MASTER_PATH)
    part_master.rename(columns={"Part Number": "Part No", "LandedCost": "Rate"}, inplace=True)

    all_data = []

    for file in input_files:
        # Read the file into a DataFrame
        df = pd.read_excel(file)
        
        # Data cleaning steps
        df = df[~df['Order Number'].str.startswith('ICPOTC')]
        df = df[df['Order Type'].isin(['Service Order', 'OTC Sales'])]
        df = df[~df['Order Item Status'].isin(['Cancelled', 'Invoiced'])]
        df['Part No'] = df['Part No'].astype(str)
        df = df[df['Qty Shipped'] > 0]
        df = pd.merge(df, location_master, on="Division", how="left")
        df['Location'] = df['Location'].fillna(df['Division'])
        df = pd.merge(df, part_master[['Part No', 'Rate', 'Category']], on="Part No", how="left")
        df['Rate'] = df['Rate'].fillna(0)
        df['Category'] = df['Category'].fillna("")
        df['Value'] = round(df['Rate'] * df['Qty Shipped'], 2)  # Rounded value
        df['Year'] = pd.DatetimeIndex(df['Date']).year
        df['Month'] = pd.to_datetime(df['Date']).dt.strftime('%b')

        required_columns = [
            'Location', 'Date', 'Division', 'Order Number', 'Order Type', 'Order Status',
            'Order Item Status', 'Type', 'Part No', 'Part Desc', 'Qty Requested',
            'Qty Shipped', 'Rate', 'Category', 'Value', 'Year', 'Month'
        ]
        df = df[required_columns]
        all_data.append(df)

    final_data = pd.concat(all_data, ignore_index=True)

    # Create the summary by Location and Order Status (dynamic types)
    summary_by_location_status = final_data.groupby(['Location', 'Order Status', 'Type'])['Value'].sum().reset_index()

    # Pivot the data so that Type values (Paid, Warranty, etc.) become rows
    summary_pivot = summary_by_location_status.pivot_table(index=['Location', 'Order Status'], columns='Type', values='Value', aggfunc='sum', fill_value=0)
    summary_pivot = summary_pivot.round(2)  # Round to 2 decimal places

    # Add a row for the Grand Total (sum across all locations and statuses)
    summary_pivot['Grand Total'] = summary_pivot.sum(axis=1)

    # Add a Grand Total column for all rows (sum across all types)
    summary_pivot.loc[('Grand Total', ''), :] = summary_pivot.sum(axis=0)

    # Create the Location monthwise summary
    summary_data = final_data.groupby(['Year', 'Month', 'Location'])['Value'].sum().reset_index()

    # Add a 'Month-Year' column for sorting and display
    summary_data['Month Number'] = summary_data['Month'].apply(lambda x: list(calendar.month_abbr).index(x[:3]))

    # Sort by Year and Month
    summary_data = summary_data.sort_values(by=['Year', 'Month'])

    # Pivot the data
    location_monthwise_pivot = summary_data.pivot_table(
        index=['Year', 'Month'], 
        columns='Location', 
        values='Value', 
        aggfunc='sum', 
        fill_value=0
    )

    # Round values to integers (remove decimals)
    location_monthwise_pivot = location_monthwise_pivot.round(0).astype(int)  # Convert to int to remove decimal points

    # Add a Grand Total column and row
    location_monthwise_pivot['Grand Total'] = location_monthwise_pivot.sum(axis=1)
    location_monthwise_pivot.loc[('Grand Total', ''), :] = location_monthwise_pivot.sum(axis=0)

    # Write all three sheets to the output Excel file
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Write the raw data sheet
        final_data.to_excel(writer, sheet_name='Data', index=False)

        # Write the summary by Order Status sheet
        summary_pivot.to_excel(writer, sheet_name='Order status wise')

        # Write the Location monthwise sheet
        location_monthwise_pivot.to_excel(writer, sheet_name='Location monthwise sheet')

    add_excel_formatting(output_path)
    print(f"Data has been cleaned, summarized, and saved to: {output_path}")

# Function to add Excel formatting
def add_excel_formatting(output_path):
    wb = load_workbook(output_path)

    # Format main data sheet
    if 'Data' in wb.sheetnames:
        data_sheet = wb['Data']
        format_sheet(data_sheet)

    # Format summary sheet
    if 'Location monthwise sheet' in wb.sheetnames:
        summary_sheet = wb['Location monthwise sheet']
        format_sheet(summary_sheet, is_summary=True, highlight_total_row=True)

    # Format order status summary sheet
    if 'Order status wise' in wb.sheetnames:
        status_sheet = wb['Order status wise']
        format_sheet(status_sheet, is_summary=True, highlight_total_row=True)

    wb.save(output_path)

def format_sheet(sheet, is_summary=False, highlight_total_row=False):
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center')

    # Apply header formatting
    for cell in sheet[1]:
        cell.fill = green_fill
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = center_alignment

    # Apply borders and center alignment to all cells
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_alignment

    # Autofit column widths
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[column].width = max_length + 2

    if highlight_total_row:
        # Find the "Grand Total" row
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == "Grand Total":
                    for grand_cell in row:
                        grand_cell.fill = green_fill
                        grand_cell.font = bold_font
                        grand_cell.alignment = center_alignment

    if is_summary:
        # Format all numerical cells as integers
        for row in sheet.iter_rows(min_row=2):  # Skip header row
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0'  # Format without decimals

# Function to open file dialog to select Location Master file
def select_location_master_file():
    global LOCATION_MASTER_PATH
    file = filedialog.askopenfilename(title="Select Location Master File", filetypes=(("Excel Files", "*.xlsx"), ("All Files", "*.*")))
    if file:
        LOCATION_MASTER_PATH = file
        location_master_file_var.set(file)

# Function to open file dialog to select Part Master file
def select_part_master_file():
    global PART_MASTER_PATH
    file = filedialog.askopenfilename(title="Select Part Master File", filetypes=(("Excel Files", "*.xlsx"), ("All Files", "*.*")))
    if file:
        PART_MASTER_PATH = file
        part_master_file_var.set(file)

# Function to open file dialog to select input folder
def select_input_folder():
    folder = filedialog.askdirectory(title="Select Input Folder")
    if folder:
        input_folder_var.set(folder)

# Function to open file dialog to select output folder
def select_output_folder():
    folder = filedialog.askdirectory(title="Select Output Folder")
    if folder:
        output_folder_var.set(folder)

# Function to generate the report (triggered by button click)
def generate_report():
    input_folder = input_folder_var.get()
    output_folder = output_folder_var.get()

    if not input_folder or not output_folder or not LOCATION_MASTER_PATH or not PART_MASTER_PATH:
        print("All folders and files need to be selected!")
        return

    input_files = [os.path.join(input_folder, file) for file in os.listdir(input_folder) if file.endswith('.xlsx')]

    if not input_files:
        print("No Excel files found in the input folder.")
        return

    output_path = get_unique_filename(output_folder, "Wip Reports.xlsx")
    clean_data(input_files, output_path)
    print(f"Report generated and saved to: {output_path}")

# Function to create a unique output file name
def get_unique_filename(folder, base_filename):
    output_path = os.path.join(folder, base_filename)
    if not os.path.exists(output_path):
        return output_path
    counter = 1
    while True:
        new_filename = f"{base_filename.split('.xlsx')[0]} ({counter}).xlsx"
        output_path = os.path.join(folder, new_filename)
        if not os.path.exists(output_path):
            return output_path
        counter += 1

# GUI Setup
root = tk.Tk()
root.title("Generate WIP Report")
root.geometry("600x500")

input_folder_var = StringVar()
output_folder_var = StringVar()
location_master_file_var = StringVar()
part_master_file_var = StringVar()

# Input folder selection
tk.Label(root, text="Select Input Folder:", font=('Arial', 10, 'bold')).pack(pady=5)
tk.Entry(root, textvariable=input_folder_var, width=50).pack(pady=5)
tk.Button(root, text="Browse", command=select_input_folder).pack(pady=5)

# Output folder selection
tk.Label(root, text="Select Output Folder:", font=('Arial', 10, 'bold')).pack(pady=5)
tk.Entry(root, textvariable=output_folder_var, width=50).pack(pady=5)
tk.Button(root, text="Browse", command=select_output_folder).pack(pady=5)

# Location Master file selection
tk.Label(root, text="Select Location Master File:", font=('Arial', 10, 'bold')).pack(pady=5)
tk.Entry(root, textvariable=location_master_file_var, width=50).pack(pady=5)
tk.Button(root, text="Browse", command=select_location_master_file).pack(pady=5)

# Part Master file selection
tk.Label(root, text="Select Part Master File:", font=('Arial', 10, 'bold')).pack(pady=5)
tk.Entry(root, textvariable=part_master_file_var, width=50).pack(pady=5)
tk.Button(root, text="Browse", command=select_part_master_file).pack(pady=5)

# Generate report button
tk.Button(root, text="Generate Report", command=generate_report, font=('Arial', 12, 'bold'), bg='green', fg='white').pack(pady=20)

root.mainloop()
