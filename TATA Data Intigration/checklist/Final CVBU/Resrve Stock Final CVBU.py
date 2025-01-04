import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

# Predefined paths for Partmaster and Location Master
PARTMASTER_PATH = r"\\tata_server\TATASERVER\TATA Data Intigration\checklist\PartmasterCVBU.xlsx"
LOCATION_MASTER_PATH = r"\\tata_server\TATASERVER\TATA Data Intigration\checklist\All Location TATA CVBU & PCBU.xlsx"

# Function to apply formatting to the Excel sheet
def apply_formatting(file_path):
    wb = load_workbook(file_path)
    
    # Format the 'Reserved Data' sheet
    sheet = wb['Reserved Data']

    # Define styles
    light_green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    bold_font = Font(bold=True)
    
    # Apply Light Green for Catg, Rate, Value, Location headers
    for col in ['A', 'B', 'C', 'D']:  # Assuming Catg, Rate, Value, Location are in columns A-D
        sheet[f'{col}1'].fill = light_green_fill
        sheet[f'{col}1'].font = bold_font
    
    # Apply Blue for other column headers starting from Part #
    other_columns = ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB']
    for col in other_columns:
        sheet[f'{col}1'].fill = blue_fill
        sheet[f'{col}1'].font = bold_font

    # Apply autofit
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    # Apply borders around the data range
    thin_border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                         top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row):
        for cell in row:
            cell.border = thin_border

    # Format the pivot table (Summary) sheet
    summary_sheet = wb['Summary']
    
    # Set the header and Grand Total formatting in light green and bold
    for cell in summary_sheet[1]:
        cell.fill = light_green_fill
        cell.font = bold_font
    
    # Format the 'Grand Total' row with bold and light green
    for row in summary_sheet.iter_rows(min_row=summary_sheet.max_row, max_row=summary_sheet.max_row):
        for cell in row:
            cell.fill = light_green_fill
            cell.font = bold_font

    # Apply autofit for the pivot table columns
    for col in summary_sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        summary_sheet.column_dimensions[column].width = adjusted_width

    # Apply borders for pivot table
    thin_border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                         top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    for row in summary_sheet.iter_rows(min_row=2, max_col=summary_sheet.max_column, max_row=summary_sheet.max_row):
        for cell in row:
            cell.border = thin_border

    # Save the changes
    wb.save(file_path)


def process_reserved_data(base_stock_files, output_dir):
    print("Starting process...")
    try:
        print("Reading Partmaster and Location Master files...")
        partmaster_data = pd.read_excel(PARTMASTER_PATH)
        location_master_data = pd.read_excel(LOCATION_MASTER_PATH)

        # Ensure required columns are present
        if 'Code' not in location_master_data.columns or 'Final Location' not in location_master_data.columns:
            messagebox.showerror("Error", "'Location Master' file must have 'Code' and 'Final Location' columns.")
            return

        # Combine all base stock files into one DataFrame
        combined_data = pd.DataFrame()

        for base_stock_file in base_stock_files:
            # Read base stock data
            base_data = pd.read_excel(base_stock_file)

            # Filter Reserved Data
            reserved_data = base_data[base_data['Availability'] == 'Reserved'].copy()

            # Ensure Part# is treated as text
            reserved_data['Part #'] = reserved_data['Part #'].astype(str)

            # Add Catg column using Partmaster VLOOKUP
            reserved_data = pd.merge(
                reserved_data,
                partmaster_data[['Part Number', 'Category']],
                left_on='Part #',
                right_on='Part Number',
                how='left'
            )
            reserved_data['Catg'] = reserved_data['Category'].fillna("")

            # Add Rate column using Partmaster VLOOKUP
            reserved_data = pd.merge(
                reserved_data,
                partmaster_data[['Part Number', 'LandedCost']],
                left_on='Part #',
                right_on='Part Number',
                how='left'
            )
            reserved_data['Rate'] = reserved_data['LandedCost'].fillna(0)

            # Add Value column (Qty * Rate)
            reserved_data['Value'] = reserved_data['Qty'] * reserved_data['Rate']

            # Add Location column using Location Master VLOOKUP
            reserved_data = pd.merge(
                reserved_data,
                location_master_data[['Code', 'Final Location']],
                left_on='Inventory Location',
                right_on='Code',
                how='left'
            )
            reserved_data['Location'] = reserved_data['Final Location'].fillna("")

            # Reorder columns and keep only the specified columns
            required_columns = ['Catg', 'Rate', 'Value', 'Location', 'Part #', 'Description', 'Qty', 
                                'Availability', 'Total Price', 'Inventory Location', 'Status', 
                                'Location 1', 'Location 2', 'Location 3', 'Min', 'Max', 'Safety', 
                                'Last Issue Date', 'TM Part Indicator', 'Product Category', 'Product Line', 
                                'Last Received Date', 'Weighted Average', 'Vendor', 'Dealer Name', 
                                'ABC Class', 'XYZ Class', 'HSN']

            # Filter only the required columns
            reserved_data = reserved_data[required_columns]

            # Append the data to the combined DataFrame
            combined_data = pd.concat([combined_data, reserved_data], ignore_index=True)

        # Create pivot table
        pivot_data = combined_data.groupby(['Location', 'Catg']).agg(
            No_of_Parts=('Catg', 'count'),  # Count number of rows
            Value=('Value', 'sum')
        ).reset_index()

        # Add grand total row to pivot
        grand_total = pd.DataFrame({
            'Location': ['Grand Total'],
            'Catg': [''],
            'No_of_Parts': [pivot_data['No_of_Parts'].sum()],
            'Value': [pivot_data['Value'].sum()]
        })
        pivot_data = pd.concat([pivot_data, grand_total], ignore_index=True)

        # Prepare output file name
        base_name = "Reserve_Stock"
        output_file = os.path.join(output_dir, f"{base_name}.xlsx")
        count = 1
        while os.path.exists(output_file):
            output_file = os.path.join(output_dir, f"{base_name}_{count}.xlsx")
            count += 1

        # Save combined data and pivot table to Excel
        with pd.ExcelWriter(output_file) as writer:
            combined_data.to_excel(writer, index=False, sheet_name='Reserved Data')
            pivot_data.to_excel(writer, index=False, sheet_name='Summary')

        # Apply formatting to the output file
        apply_formatting(output_file)

        print(f"File saved: {output_file}")
        messagebox.showinfo("Success", f"All files processed successfully! Output saved in: {output_dir}")

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")


def open_files_and_output():
    # Open file dialog to select multiple base stock files
    base_stock_files = filedialog.askopenfilenames(title="Select Base Stock files", filetypes=[("Excel Files", "*.xlsx")])

    if not base_stock_files:
        messagebox.showerror("Error", "No base stock files selected.")
        return

    # Open file dialog to select output directory
    output_dir = filedialog.askdirectory(title="Select Output Directory")

    if not output_dir:
        messagebox.showerror("Error", "No output directory selected.")
        return

    # Process the data from the selected files
    process_reserved_data(base_stock_files, output_dir)


# Set up GUI
root = tk.Tk()
root.title("Reserved Data Processor")
root.geometry("500x300")  # Set window size

# Create and place the labels
file_label = tk.Label(root, text="Select Base Stock Files:")
file_label.pack(pady=10)

select_button = tk.Button(root, text="Select Files", command=open_files_and_output)
select_button.pack(pady=10)

output_label = tk.Label(root, text="Select Output Directory:")
output_label.pack(pady=10)

run_button = tk.Button(root, text="Run Processing", command=lambda: open_files_and_output())
run_button.pack(pady=30)

root.mainloop()
