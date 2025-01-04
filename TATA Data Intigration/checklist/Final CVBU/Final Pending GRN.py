import os
import pandas as pd
from tkinter import Tk, filedialog, messagebox, StringVar, Label, Button
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# Define columns to include
NEW_COLUMNS = ["Location", "Vendor", "Month", "Year", "Casual/VOR"]
COLUMN_ORDER = [
    "Division Name", "Order #", "Commit Flag", "Spares Order Type", "Part #", "Status", "Ware House Name",
    "SAP Invoice #", "Invoice_Date", "Recd Qty", "Line Item Invoice Total", "Total_Invoice_Amount",
    "Discount Amount", "Other Charges Amount", "Net Amount", "VAT", "CST VAT", "CST", "CST Surcharge",
    "LST", "LST Surcharge", "Additional Tax", "TOT", "Octroi", "Vendor Name", "Payer Code", "SAP Order Num",
    "IRN", "IRN Date", "Weighted Avg", "Cash Discount", "Cash Discount Percentage", "Discount Per Part",
    "Discount Per Part Percentage", "CGST", "IGST", "SGST", "UTGST", "GST Invoice #", "TCS Amount",
    "Challan #", "Transaction Date", "Transaction Number", "Order Type", "Challan Date", "Challan Quantity",
    "Purchase_Order_Date", "Movement Type", "Condition", "Vendor Invoice #"
]

LOCATION_MAPPING_FILE = r"\\tata_server\TATASERVER\TATA Data Intigration\checklist\All Location TATA CVBU & PCBU.xlsx"

def process_pending_grn(input_folder, output_folder):
    try:
        # Collect all Excel files in the input folder
        file_paths = [os.path.join(input_folder, file) for file in os.listdir(input_folder) if file.endswith('.xlsx')]
        if not file_paths:
            messagebox.showwarning("Warning", "No Excel files found in the selected folder.")
            return

        # Read all files into DataFrames, skipping empty or all-NA ones
        dfs = []
        for file in file_paths:
            df = pd.read_excel(file, engine='openpyxl')
            if not df.empty and not df.isna().all(axis=None):  # Exclude empty and all-NA DataFrames
                dfs.append(df)

        if not dfs:  # If no valid DataFrames are found
            messagebox.showinfo("Info", "No valid data found in the provided files.")
            return

        # Concatenate the valid DataFrames
        compiled_df = pd.concat(dfs, ignore_index=True)

        # Filter for Pending GRN data
        if 'Status' not in compiled_df.columns:
            messagebox.showerror("Error", "The files must contain a 'Status' column.")
            return

        pending_grn_df = compiled_df[compiled_df['Status'] == 'In Transit']

        if pending_grn_df.empty:
            messagebox.showinfo("Info", "No Pending GRN records found.")
            return

        # Add new columns with formulas or mappings
        pending_grn_df = add_calculated_columns(pending_grn_df)

        # Replace all occurrences of 'Rs.' and handle float conversion
        pending_grn_df = replace_rs_and_convert_to_float(pending_grn_df)

        # Reorder columns with new ones at the start
        reordered_columns = NEW_COLUMNS + [col for col in COLUMN_ORDER if col in pending_grn_df.columns]
        pending_grn_df = pending_grn_df[reordered_columns]

        # Get a unique file name
        output_file = get_unique_filename(output_folder, "Pending_GRN")

        # Save the filtered data to an Excel file with styled headers
        save_and_format(output_file, pending_grn_df)

        # Create Pivot Table for Combined Summary
        create_combined_summary_pivot(output_file, pending_grn_df)

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

def add_calculated_columns(df):
    """Add the new columns with formulas or mapped data."""
    # Load location mapping data
    location_mapping = pd.read_excel(LOCATION_MAPPING_FILE, sheet_name="Sheet1")
    location_dict = dict(zip(location_mapping.iloc[:, 0], location_mapping.iloc[:, 1]))

    # Perform VLOOKUP-like mapping for Location
    df["Location"] = df["Division Name"].map(location_dict).fillna(df["Division Name"])

    # Classify Vendor based on the first five characters of "Order #"
    df["Vendor"] = df["Order #"].apply(
        lambda x: "Non TATA" if str(x).startswith(("CPPUR", "ECPUR", "ICPPUR", "EICPUR")) else "TATA Motors"
    )

    # Format Month and Year from Invoice Date
    df["Month"] = pd.to_datetime(df["Invoice_Date"], errors="coerce").dt.strftime("%b-%y").fillna("")
    df["Year"] = pd.to_datetime(df["Invoice_Date"], errors="coerce").dt.year.fillna("")

    # Add Casual/VOR based on Spares Order Type
    df["Casual/VOR"] = df["Spares Order Type"].apply(
        lambda x: "VOR Purchase" if any(keyword in str(x) for keyword in ["VOR Order PVBU","VOR Order CVBU", "CP-VOR Order CVBU"]) else 
                  ("Casual Purchase" if any(keyword in str(x) for keyword in ["Casual Order pcbu", "Casual Order PVBU", "Lub Order"]) else 
                   "Casual Purchase")  # Default to Casual Purchasee if no match
    )

    return df

def replace_rs_and_convert_to_float(df):
    """Replace all occurrences of 'Rs.' and convert columns to float where applicable."""
    # Remove 'Rs.' and any non-numeric characters, and convert columns to numeric
    df = df.applymap(lambda x: str(x).replace('Rs.', '').replace(',', '') if isinstance(x, str) else x)

    # Convert columns that should be numeric to float
    numeric_columns = [
        "Line Item Invoice Total", "Total_Invoice_Amount", "Discount Amount", "Other Charges Amount",
        "Net Amount", "VAT", "CST VAT", "CST", "CST Surcharge", "LST", "LST Surcharge", "Additional Tax", 
        "TOT", "Octroi", "Weighted Avg", "Cash Discount", "Cash Discount Percentage", "Discount Per Part",
        "Discount Per Part Percentage", "CGST", "IGST", "SGST", "UTGST", "TCS Amount"
    ]

    for col in numeric_columns:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')  # Convert to numeric, coercing errors to NaN

    return df

def save_and_format(file_path, df):
    try:
        df.to_excel(file_path, index=False, engine='openpyxl')
        wb = load_workbook(file_path)
        ws = wb.active

        # Set column widths and apply header styles
        light_green = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        light_blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.cell(row=1, column=col_idx).value = col_name

            if col_name in NEW_COLUMNS:
                ws.cell(row=1, column=col_idx).fill = light_green
            else:
                ws.cell(row=1, column=col_idx).fill = light_blue

            max_length = max((len(str(cell.value)) for cell in ws[col_idx] if cell.value), default=10)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_length + 2

        border_style = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border_style
                cell.alignment = Alignment(horizontal='center', vertical='center')

        wb.save(file_path)
        messagebox.showinfo("Success", f"Pending GRN report saved to: {file_path}")
    except Exception as e:
        messagebox.showerror("Error", f"Error saving file {file_path}: {e}")

def create_combined_summary_pivot(file_path, df):
    try:
        wb = load_workbook(file_path)
        summary_sheet = wb.create_sheet(title="Combined Summary")

        # Ensure 'Year' and 'Month' are strings and handle missing values (NaN)
        df['Year'] = df['Year'].apply(str).fillna('')
        df['Month'] = df['Month'].apply(str).fillna('')

        # Replace "Non TATA" with "TGP FROM CODEALER" in the 'Vendor' column
        df["Vendor"] = df["Vendor"].replace("Non TATA", "TGP FROM CODEALER")
        df["Vendor"] = df["Vendor"].replace("TATA Motors", "TGP FROM TATA")

        # Pivot Table for Combined Summary: Summing Line Item Invoice Total
        pivot_df = pd.pivot_table(
            df,
            index=["Vendor", "Casual/VOR", "Location"],
            columns=["Month"],
            values="Line Item Invoice Total",
            aggfunc="sum",
            fill_value=0
        ).reset_index()

        # Round float values to integers
        pivot_df = pivot_df.round(0)

        # Ensure the months are in the desired chronological order
        month_order = ["Jan-24", "Feb-24", "Mar-24", "Apr-24", "May-24", "Jun-24", "Jul-24", "Aug-24", "Sep-24", "Oct-24", "Nov-24", "Dec-24"]
        sorted_months = [col for col in month_order if col in pivot_df.columns]
        static_columns = [col for col in ["Location", "Vendor", "Casual/VOR"] if col in pivot_df.columns]
        pivot_df = pivot_df[static_columns + sorted_months]

        # Remove month columns where the total is 0
        non_zero_months = [col for col in sorted_months if pivot_df[col].sum() != 0]
        pivot_df = pivot_df[static_columns + non_zero_months]

        # Sort by 'Location' in ascending order
        pivot_df = pivot_df.sort_values(by=["Location"], ascending=True)
        
        # Add a "Total" column for row-wise sum of all monthly columns
        pivot_df["Total"] = pivot_df.iloc[:, len(static_columns):].sum(axis=1)

        # Calculate Grand Total row
        grand_total = pivot_df.iloc[:, len(static_columns):].sum(axis=0)
        grand_total["Location"] = "Grand Total"
        grand_total["Vendor"] = ""
        grand_total["Casual/VOR"] = ""
        pivot_df = pd.concat([pivot_df, pd.DataFrame([grand_total])], ignore_index=True)

        # Write the pivot table to the new sheet
        for row in dataframe_to_rows(pivot_df, index=False, header=True):
            summary_sheet.append(row)

        # Apply styles: Center alignment and borders
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), 
            right=Side(style="thin"),
            top=Side(style="thin"), 
            bottom=Side(style="thin")
        )

        for row in summary_sheet.iter_rows():
            for cell in row:
                cell.alignment = center_alignment
                cell.border = thin_border

        # Apply styles to header row
        header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light Blue
        bold_font = Font(bold=True)

        for cell in summary_sheet[1]:
            cell.fill = header_fill
            cell.font = bold_font

        # Apply Grand Total row formatting (same as the header)
        grand_total_row = summary_sheet.max_row
        for cell in summary_sheet[grand_total_row]:
            cell.fill = header_fill
            cell.font = bold_font

        # Bold the "Total" column
        total_col_idx = summary_sheet.max_column
        for row_idx in range(2, grand_total_row):  # Exclude header row
            cell = summary_sheet.cell(row=row_idx, column=total_col_idx)
            cell.font = Font(bold=True)

        # Format cells to remove decimals
        for row in summary_sheet.iter_rows(min_row=2, max_row=summary_sheet.max_row, min_col=4, max_col=summary_sheet.max_column):
            for cell in row:
                if isinstance(cell.value, (int, float)):  # Check if the value is numeric
                    cell.number_format = '0'  # Format as integer (no decimals)

        # Auto-fit column widths
        for col_idx in range(1, summary_sheet.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in summary_sheet[col_letter]:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            summary_sheet.column_dimensions[col_letter].width = max_length + 2

        # Save the file with the pivot table and formatting
        wb.save(file_path)
        messagebox.showinfo("Success", f"Combined Summary pivot table saved to: {file_path}")

    except Exception as e:
        messagebox.showerror("Error", f"Error generating pivot table: {e}")

def get_unique_filename(folder_path, base_name):
    """Generate a unique filename to prevent overwriting."""
    counter = 0
    while True:
        file_name = f"{base_name}_{counter}.xlsx" if counter > 0 else f"{base_name}.xlsx"
        full_path = os.path.join(folder_path, file_name)
        if not os.path.exists(full_path):
            return full_path
        counter += 1

# GUI Section
def select_input_folder():
    """Allow user to select the input folder."""
    folder = filedialog.askdirectory(title="Select Input Folder")
    if folder:
        input_folder_var.set(folder)

def select_output_folder():
    """Allow user to select the output folder."""
    folder = filedialog.askdirectory(title="Select Output Folder")
    if folder:
        output_folder_var.set(folder)

def generate_report():
    """Trigger the Pending GRN report generation."""
    input_folder = input_folder_var.get()
    output_folder = output_folder_var.get()

    if not input_folder:
        messagebox.showerror("Error", "Please select an input folder.")
        return

    if not output_folder:
        messagebox.showerror("Error", "Please select an output folder.")
        return

    process_pending_grn(input_folder, output_folder)

# Main Application Window
root = Tk()
root.title("Pending GRN Report Generator")
root.geometry("600x300")
root.resizable(False, False)

# Variables to store folder paths
input_folder_var = StringVar()
output_folder_var = StringVar()

# GUI Layout
Label(root, text="Select Input Folder:", font=('Arial', 10, 'bold')).pack(pady=10, anchor='w', padx=20)
Button(root, text="Browse", command=select_input_folder, width=20).pack(pady=5)
Label(root, textvariable=input_folder_var, wraplength=500).pack(anchor='w', padx=20)

Label(root, text="Select Output Folder:", font=('Arial', 10, 'bold')).pack(pady=10, anchor='w', padx=20)
Button(root, text="Browse", command=select_output_folder, width=20).pack(pady=5)
Label(root, textvariable=output_folder_var, wraplength=500).pack(anchor='w', padx=20)

Button(root, text="Generate Pending GRN Report", command=generate_report, width=30, font=('Arial', 12)).pack(pady=20)

# Run the GUI event loop
root.mainloop()
