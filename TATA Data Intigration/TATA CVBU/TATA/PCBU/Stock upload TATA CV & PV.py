import os
import pandas as pd
from tkinter import Tk, Label, Button, filedialog, messagebox
import tkinter as tk
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from tkinter.ttk import Progressbar
import logging
import threading

logging.basicConfig(filename='error.log', level=logging.ERROR)

def load_location_mapping(mapping_file):
    try:
        mapping_df = pd.read_excel(mapping_file, engine='openpyxl')
        if 'Code' not in mapping_df.columns or 'Final Location' not in mapping_df.columns:
            raise ValueError("Mapping file must contain 'Code' and 'Final Location' columns.")
        return dict(zip(mapping_df['Code'], mapping_df['Final Location']))
    except Exception as e:
        messagebox.showerror("Error", f"Error loading location mapping file: {e}")
        return {}

def compile_and_process_excel_files(folder_path, output_file, mapping_file):
    location_mapping = load_location_mapping(mapping_file)
    try:
        files = os.listdir(folder_path)
    except (FileNotFoundError, PermissionError) as e:
        messagebox.showerror("Error", f"{e}")
        return

    excel_files = [f for f in files if f.endswith('.xlsx')]

    if not excel_files:
        messagebox.showinfo("Info", "No Excel files found in the folder.")
        return

    compiled_df = pd.concat([pd.read_excel(os.path.join(folder_path, file)) for file in excel_files], ignore_index=True)

    required_columns = ['Part #', 'Qty', 'Inventory Location', 'Availability', 'Status']
    if not all(col in compiled_df.columns for col in required_columns):
        messagebox.showerror("Error", "Missing columns in DataFrame")
        return

    filtered_df = compiled_df[(compiled_df['Availability'] == 'On Hand') & (compiled_df['Status'] == 'Good')]
    filtered_df = filtered_df[['Part #', 'Qty', 'Inventory Location']]
    filtered_df = filtered_df.rename(columns={'Part #': 'Partnumber', 'Inventory Location': 'Location'})
    filtered_df['Partnumber'] = filtered_df['Partnumber'].astype(str)
    filtered_df['Location'] = filtered_df['Location'].map(lambda x: location_mapping.get(x, x))

    try:
        filtered_df.to_excel(output_file, index=False, engine='openpyxl')
        wb = load_workbook(output_file)
        ws = wb.active

        # Debugging print to check columns
        columns_in_worksheet = [col[0].value for col in ws.iter_cols(1, ws.max_column)]
        print(f"Columns in worksheet: {columns_in_worksheet}")  # Debug output for column headers

        partnumber_column = next((col[0].column_letter for col in ws.iter_cols(1, ws.max_column) if col[0].value == 'Partnumber'), None)

        if partnumber_column:
            print(f"Partnumber column: {partnumber_column}")  # Debug to confirm column found
            for cell in ws[partnumber_column]:
                cell.number_format = '@'  # Set number format as text
        else:
            print("Partnumber column not found")  # Debug output if the column is missing

        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

        for cell in ws[1]:
            cell.font = Font(bold=False)

        wb.save(output_file)
        messagebox.showinfo("Success", f"Processed Excel file '{output_file}' successfully.")
    except Exception as e:
        logging.error(f"Error saving file {output_file}: {e}")
        messagebox.showerror("Error", f"Error saving file {output_file}: {e}")

def select_folder():
    folder_selected = filedialog.askdirectory(initialdir=folder_path_var.get() or 'C:\\')
    if folder_selected:
        folder_path_var.set(folder_selected)

def select_file():
    file_selected = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], initialfile="stock upload.xlsx")
    if file_selected:
        output_file_var.set(file_selected)

def select_mapping_file():
    file_selected = filedialog.askopenfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if file_selected:
        mapping_file_var.set(file_selected)

def process_files():
    folder_path, output_file, mapping_file = folder_path_var.get(), output_file_var.get(), mapping_file_var.get()
    if not all([folder_path, output_file, mapping_file]):
        messagebox.showwarning("Warning", "Please select folder, output file, and mapping file.")
        return
    compile_and_process_excel_files(folder_path, output_file, mapping_file)

def process_files_async():
    threading.Thread(target=process_files).start()

root = Tk()
root.title("Excel File Compiler")

folder_path_var = tk.StringVar()
output_file_var = tk.StringVar()
mapping_file_var = tk.StringVar()

Label(root, text="Select Folder:").grid(row=0, column=0, padx=5, pady=5)
Button(root, text="Browse", command=select_folder).grid(row=0, column=1, padx=5, pady=5)
Label(root, textvariable=folder_path_var).grid(row=0, column=2, padx=5, pady=5)

Label(root, text="Select Output File:").grid(row=1, column=0, padx=5, pady=5)
Button(root, text="Browse", command=select_file).grid(row=1, column=1, padx=5, pady=5)
Label(root, textvariable=output_file_var).grid(row=1, column=2, padx=5, pady=5)

Label(root, text="Select Mapping File:").grid(row=2, column=0, padx=5, pady=5)
Button(root, text="Browse", command=select_mapping_file).grid(row=2, column=1, padx=5, pady=5)
Label(root, textvariable=mapping_file_var).grid(row=2, column=2, padx=5, pady=5)

Button(root, text="Process Files", command=process_files_async).grid(row=3, column=0, columnspan=3, pady=20)

progress = Progressbar(root, orient="horizontal", length=300, mode="determinate")
progress.grid(row=4, column=0, columnspan=3, pady=10)

status_var = tk.StringVar()
Label(root, textvariable=status_var).grid(row=5, column=0, columnspan=3, pady=5)

root.mainloop()
