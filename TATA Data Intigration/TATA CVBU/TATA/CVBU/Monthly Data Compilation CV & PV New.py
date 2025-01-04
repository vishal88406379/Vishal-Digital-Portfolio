import os
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

def log_error(message):
    """Logs errors to an error log file."""
    with open("error_log.txt", "a") as f:
        f.write(f"{pd.Timestamp.now()}: {message}\n")

def compile_excel_files(folder_path, output_file):
    """Compiles all Excel files from a folder into a single output file."""
    try:
        files = os.listdir(folder_path)
    except FileNotFoundError:
        log_error(f"The folder path '{folder_path}' does not exist.")
        return "Folder path does not exist."
    except PermissionError:
        log_error(f"Permission denied for folder path '{folder_path}'.")
        return "Permission denied to access the folder."

    excel_files = [f for f in files if f.endswith('.xlsx')]

    if not excel_files:
        return "No Excel files found in the folder."

    compiled_df = pd.DataFrame()

    for file in excel_files:
        input_file = os.path.join(folder_path, file)

        try:
            # Read each Excel file into a DataFrame without modifying any data
            df = pd.read_excel(input_file, sheet_name=None)  # Read all sheets
            # Combine all sheets in each Excel file
            for sheet_name, sheet_data in df.items():
                compiled_df = pd.concat([compiled_df, sheet_data], ignore_index=True)
        except Exception as e:
            log_error(f"Error processing file {file}: {e}")
            return f"Error processing file {file}: {e}"

    try:
        wb = Workbook()
        ws = wb.active

        # Write the DataFrame to the output Excel file without modifying its structure
        for r in dataframe_to_rows(compiled_df, index=False, header=True):
            ws.append(r)

        # Adjust column widths based on content
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Make the header bold
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Save the file
        wb.save(output_file)
        return f"File saved successfully as {output_file}"

    except Exception as e:
        log_error(f"Error saving file: {e}")
        return f"Error saving file: {e}"

def browse_folder(radio_value):
    """Allows user to select a folder for input files."""
    folder_path = filedialog.askdirectory()
    folder_entries[radio_value].delete(0, tk.END)
    folder_entries[radio_value].insert(0, folder_path)

def browse_output_folder():
    """Allows user to select a folder for output files."""
    folder_path = filedialog.askdirectory()
    output_folder_entry.delete(0, tk.END)
    output_folder_entry.insert(0, folder_path)

def run_compile():
    """Runs the compilation process for all selected keywords and folders."""
    output_folder = output_folder_entry.get()
    if not output_folder:
        messagebox.showwarning("TATA SCS Monthly Data", "Please select an output folder.")
        return

    results = []

    try:
        # Process all keywords
        for idx, keyword in enumerate(keywords):
            folder_path = folder_entries[idx].get()
            if folder_path:
                output_file = os.path.join(output_folder, f"{keyword}.xlsx")
                result = compile_excel_files(folder_path, output_file)
                results.append(f"{keyword}: {result}")
            else:
                results.append(f"{keyword}: Skipped due to missing folder.")

        # Display final message box with all results
        final_message = "\n".join(results)
        messagebox.showinfo("TATA SCS Monthly Data", f"Processing complete:\n\n{final_message}")

    finally:
        # Optionally open the output folder
        os.startfile(output_folder)  # For Windows

# Create the main window
root = tk.Tk()
root.title("TATA SCS Excel File Compiler")

# Create frames for better layout
frame_header = tk.Frame(root, padx=10, pady=10)
frame_header.pack(fill=tk.X)

frame_inputs = tk.Frame(root, padx=10, pady=10)
frame_inputs.pack(fill=tk.BOTH, expand=True)

frame_buttons = tk.Frame(root, padx=10, pady=10)
frame_buttons.pack(fill=tk.X)

# Header Label
header_label = tk.Label(frame_header, text="TATA SCS Monthly Data Compiler", font=("Arial", 16, "bold"))
header_label.pack()

# Keywords selection and control layout
keywords = [
    "OTC INVOICE", "PURCHASE LINE PO", "Purchase Line Items",
    "SPARE CONSUMPTION", "CLOSING STOCK", "stock transaction",
    "channel partner", "Job Line Invoice", "sap purchase order reason"
]

folder_entries = {}
output_folder_entry = tk.Entry(frame_buttons, width=40)
output_folder_entry.pack(padx=5, pady=5)

# Output folder path button
browse_output_folder_button = tk.Button(frame_buttons, text="Browse Output Folder", command=browse_output_folder, bg="lightgreen")
browse_output_folder_button.pack()

run_buttons = []
for idx, keyword in enumerate(keywords):
    tk.Label(frame_inputs, text=f"{keyword}:", font=("Arial", 12, "bold")).grid(row=idx, column=0, padx=5, pady=5, sticky="w")

    tk.Label(frame_inputs, text="Folder:", font=("Arial", 10)).grid(row=idx, column=1, padx=5, pady=5, sticky="w")
    folder_entry = tk.Entry(frame_inputs, width=40)
    folder_entry.grid(row=idx, column=2, padx=5, pady=5, sticky="ew")
    folder_entries[idx] = folder_entry
    browse_folder_button = tk.Button(frame_inputs, text="Browse Folder", command=lambda v=idx: browse_folder(v), bg="lightblue")
    browse_folder_button.grid(row=idx, column=3, padx=5, pady=5)

# Run All button
run_all_button = tk.Button(frame_buttons, text="Run All", command=run_compile, bg="blue", fg="white", font=('Arial', 12, 'bold'))
run_all_button.pack()

# Configure column weights for expanding
frame_inputs.grid_columnconfigure(2, weight=1)

# Start the GUI event loop
root.mainloop()
