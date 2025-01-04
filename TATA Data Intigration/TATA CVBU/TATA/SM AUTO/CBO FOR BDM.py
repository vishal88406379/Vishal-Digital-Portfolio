import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl.utils import get_column_letter
from datetime import datetime

def load_input_file():
    input_file_path.set(filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")]))

def load_output_file():
    output_file_path.set(filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                       filetypes=[("Excel files", "*.xlsx;*.xls")]))

def load_location_file():
    location_file_path.set(filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")]))

def process_files():
    try:
        input_path = input_file_path.get()
        output_path = output_file_path.get()
        location_path = location_file_path.get()

        if not input_path or not output_path or not location_path:
            messagebox.showwarning("Warning", "Please select all files.")
            return

        # Load input file
        df = pd.read_excel(input_path)

        # Print column names for debugging
        print("Columns in input file:", df.columns.tolist())

        # Rename columns
        df.rename(columns={'Unnamed: 1': 'Party Name', 'Unnamed: 6': 'Description'}, inplace=True)

        # Convert 'Product' column to string
        if 'Product' in df.columns:
            df['Product'] = df['Product'].astype(str)
        else:
            print("Column 'Product' not found in input data.")

        # Check if 'ZShip From' exists
        if 'ZShip From' not in df.columns:
            raise KeyError("'ZShip From' column not found in input file.")

        # Convert 'Created On' to datetime and extract month
        if 'Created On' in df.columns:
            df['Created On'] = pd.to_datetime(df['Created On'], dayfirst=True, errors='coerce')
            df['Days'] = (datetime.today() - df['Created On']).dt.days  # Calculate difference in days
            
            # Define cutoff date as June 26, 2024
            cutoff_date = datetime(2024, 6, 26)
            
            # Remove rows where 'Created On' is on or before the cutoff date
            df = df[df['Created On'] > cutoff_date]
        else:
            print("Column 'Created On' not found in input data.")

        # Filter out specific Party Names that start with "CSM" or are "Tata Motors Ltd."
        df = df[~df['Party Name'].str.startswith('CSM') & ~df['Party Name'].isin(['Tata Motors Ltd.'])]

        # Print first few rows for debugging
        print("First few rows of input data after processing:")
        print(df.head())

        # Remove 'ea' and convert to float then to int for specified columns
        quantity_columns = [
            'Requested Quantity', 'Confirmed Quantity', 
            'Fulfilled Quantity', 'Invoiced Quantity', 
            'Pending Qty'
        ]

        for col in quantity_columns:
            if col in df.columns:
                print(f"Processing column: {col}")
                df[col] = df[col].str.replace('ea', '').str.replace(',', '').astype(float).fillna(0).astype(int)
            else:
                print(f"Column '{col}' not found in input data.")

        # Load location mapping from 'Monthly data locations' sheet
        try:
            location_mapping_df = pd.read_excel(location_path, sheet_name='Monthly data locations')
        except ValueError as e:
            messagebox.showerror("Error", "Sheet 'Monthly data locations' not found. Please check the sheet name.")
            print("Available sheets:", pd.ExcelFile(location_path).sheet_names)
            return

        print("Columns in location mapping file:", location_mapping_df.columns.tolist())

        # Check if 'ZShip From' exists in mapping
        if 'ZShip From' not in location_mapping_df.columns or 'Location Name' not in location_mapping_df.columns:
            raise KeyError("'ZShip From' or 'Location Name' column not found in location mapping file.")

        # Create location mapping dictionary
        location_mapping = dict(zip(location_mapping_df['ZShip From'], location_mapping_df['Location Name']))

        # Add Location name based on mapping
        df['Location name'] = df['ZShip From'].map(location_mapping)

        # Define the desired column order
        desired_columns = [
            'Account', 'Party Name', 'Sales Order', 'External Partner',
            'Created On', 'Days', 'Product', 'Description', 'Location name',
            'ZShip From', 'External Reference', 'Requested Quantity',
            'Confirmed Quantity', 'Fulfilled Quantity', 'Invoiced Quantity',
            'Pending Qty',
        ]

        # Ensure the DataFrame has all the desired columns
        for col in desired_columns:
            if col not in df.columns:
                df[col] = pd.NA  # Add the column with NA values if it doesn't exist

        # Reorder DataFrame
        df = df[desired_columns]

        # Save output with formatting
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Processed Data')
            workbook = writer.book
            worksheet = writer.sheets['Processed Data']

            # Apply text format to 'Product' column
            product_col_index = df.columns.get_loc('Product') + 1  # +1 for 1-based indexing in Excel
            for row in range(2, len(df) + 2):  # Start from row 2 for data
                worksheet.cell(row=row, column=product_col_index).number_format = '@'  # Set format to text

            # Apply date format to 'Created On' column
            created_on_col_index = df.columns.get_loc('Created On') + 1
            for row in range(2, len(df) + 2):
                worksheet.cell(row=row, column=created_on_col_index).number_format = 'DD-MM-YYYY'  # Set format to dd-mm-yyyy

            # Apply text format to 'Days' column
            days_col_index = df.columns.get_loc('Days') + 1
            for row in range(2, len(df) + 2):
                worksheet.cell(row=row, column=days_col_index).number_format = '0'  # Set format to integer

            # Autofit columns
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        messagebox.showinfo("Success", "File processed and saved successfully! The 'Created On' date is formatted as 'dd-mm-yyyy' and 'Days' column added.")

    except Exception as e:
        messagebox.showerror("Error", str(e))
        print("Error details:", str(e))

def create_gui():
    global input_file_path, output_file_path, location_file_path

    root = tk.Tk()
    root.title("Excel File Processor")

    input_file_path = tk.StringVar()
    output_file_path = tk.StringVar()
    location_file_path = tk.StringVar()

    tk.Label(root, text="Select Input Excel File:").pack(pady=5)
    tk.Entry(root, textvariable=input_file_path, width=50).pack(pady=5)
    tk.Button(root, text="Browse", command=load_input_file).pack(pady=5)

    tk.Label(root, text="Select Output Path:").pack(pady=5)
    tk.Entry(root, textvariable=output_file_path, width=50).pack(pady=5)
    tk.Button(root, text="Browse", command=load_output_file).pack(pady=5)

    tk.Label(root, text="Select Location Mapping File:").pack(pady=5)
    tk.Entry(root, textvariable=location_file_path, width=50).pack(pady=5)
    tk.Button(root, text="Browse", command=load_location_file).pack(pady=5)

    tk.Button(root, text="Run", command=process_files).pack(pady=20)

    root.mainloop()

if __name__ == "__main__":
    create_gui()
