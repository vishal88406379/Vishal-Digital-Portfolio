import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
import traceback

class ExcelMapperApp:
    def __init__(self, master):
        self.master = master
        self.master.title("Excel Location Mapper")
        self.master.geometry("400x200")

        self.input_files = []
        self.output_dir = ""

        self.create_widgets()

    def create_widgets(self):
        tk.Button(self.master, text="Select Excel Files", command=self.select_input_files).pack(pady=10)
        tk.Button(self.master, text="Select Output Directory", command=self.select_output_directory).pack(pady=10)
        tk.Button(self.master, text="Run Mapping", command=self.replace_values).pack(pady=20)

    def select_input_files(self):
        self.input_files = filedialog.askopenfilenames(
            title="Select Excel Files",
            filetypes=[("Excel files", "*.xlsx;*.xls")]
        )
        print(f"Selected input files: {self.input_files}")  # Debug statement

    def select_output_directory(self):
        self.output_dir = filedialog.askdirectory(title="Select Output Directory")
        print(f"Selected output directory: {self.output_dir}")  # Debug statement

    def replace_values(self):
        if not self.input_files or not self.output_dir:
            messagebox.showerror("Error", "Please select input files and output directory.")
            return

        error_messages = []
        
        for input_file in self.input_files:
            try:
                self.process_file(input_file)
            except Exception as e:
                error_messages.append(f"Error processing {input_file}: {str(e)}")
                print(traceback.format_exc())

        if error_messages:
            messagebox.showerror("Errors Encountered", "\n".join(error_messages))
        else:
            messagebox.showinfo("Success", "All files processed and saved successfully!")

    def process_file(self, input_file):
        df = pd.read_excel(input_file)

        if 'Site' not in df.columns:
            raise ValueError(f"'Site' column is missing from the file: {input_file}")

        # Map Site to Location
        site_location_mapping = self.get_site_location_mapping()
        df['Location'] = df['Site'].map(site_location_mapping)

        # Clean and convert quantity columns
        self.clean_and_convert_qty(df)

        # Save modified DataFrame with the new file name
        output_file = f"{self.output_dir}/Stock Sm Auto Group All locations.xlsx"
        df.to_excel(output_file, index=False, engine='openpyxl')

        # Create files for specified locations
        self.save_location_files(df)

        # Adjust formats in the saved file
        self.adjust_formats(output_file)

        print(f"File processed and saved successfully: {output_file}")

    def save_location_files(self, df):
        # Define locations for Retail and Tass
        retail_locations = [
            'Varanasi Retail',
            'Lucknow STU Retail',
            'Lucknow Retail',
            'Gorakhpur Retail',
            'Allahabad Retail',
            'Chopan Retail',
            'Faizabad Retail',
            'Fatehpur Retail'
        ]

        tass_locations = [
            'Varanasi Tass',
            'Lucknowes TASS'
        ]

        # Filter for retail locations and remove rows with Ending Qty <= 0
        retail_df = df[(df['Location'].isin(retail_locations)) & (df['Ending Qty'] > 0)]
        retail_df = self.format_output_dataframe(retail_df)

        retail_file = f"{self.output_dir}/Stock upload Retail.xlsx"
        retail_df.to_excel(retail_file, index=False, engine='openpyxl')
        self.adjust_formats(retail_file)  # Adjust formats for retail file
        print(f"Retail locations file created: {retail_file}")

        # Filter for tass locations and remove rows with Ending Qty <= 0
        tass_df = df[(df['Location'].isin(tass_locations)) & (df['Ending Qty'] > 0)]
        tass_df = self.format_output_dataframe(tass_df)

        tass_file = f"{self.output_dir}/Stock upload Tass.xlsx"
        tass_df.to_excel(tass_file, index=False, engine='openpyxl')
        self.adjust_formats(tass_file)  # Adjust formats for tass file
        print(f"Tass locations file created: {tass_file}")

    def format_output_dataframe(self, df):
        """ Format the DataFrame to match required output structure. """
        output_df = pd.DataFrame({
            'Partnumber': df['Product'].astype(str),  # Ensuring Partnumber is treated as text
            'Qty': df.get('Ending Qty', pd.Series(dtype='float')),  # Maintain Ending Qty as Qty
            'Location': df['Location']  # Include Location
        })

        return output_df

    def get_site_location_mapping(self):
        return {
            'C1.2VNS.RET': 'Varanasi Retail',
            'C1.2VNS.SER': 'Varanasi Tass',
            'C1.3LKO.INS': 'Lucknow STU Retail',
            'C1.3LKO.RET': 'Lucknow Retail',
            'C1.3LKO.SER': 'Lucknowes TASS',
            'C1.3LKO.VOR': 'Lucknowvor Retail',
            'C1.4GKP.RET': 'Gorakhpur Retail',
            'C1.5ALD.RET': 'Allahabad Retail',
            'C1.6SON.RET': 'Chopan Retail',
            'C1.7FZP.RET': 'Faizabad Retail',
            'C1.9FTP.RET': 'Fatehpur Retail'
        }

    def clean_and_convert_qty(self, df):
        for qty_col in ['Opening Qty', 'Ending Qty']:
            if qty_col in df.columns:
                df[qty_col] = pd.to_numeric(
                    df[qty_col].replace({'ea': '', ',': ''}, regex=True),
                    errors='coerce'
                )
            else:
                raise ValueError(f"'{qty_col}' column is missing from the DataFrame.")

    def adjust_formats(self, output_file):
        wb = load_workbook(output_file)
        ws = wb.active

        # Apply Text format to Partnumber
        partnumber_col_idx = self.get_column_index(ws, 'Partnumber')
        if partnumber_col_idx:
            for cell in ws[ws.cell(row=1, column=partnumber_col_idx).column_letter]:
                cell.number_format = '@'  # Set format to Text

        # Apply General format to quantity columns
        for col_name in ['Opening Qty', 'Ending Qty']:
            col_idx = self.get_column_index(ws, col_name)
            if col_idx:
                for cell in ws[ws.cell(row=1, column=col_idx).column_letter]:
                    cell.number_format = 'General'

        # Auto-fit columns
        self.auto_fit_columns(ws)

        wb.save(output_file)

    def get_column_index(self, ws, col_name):
        for idx, cell in enumerate(ws[1], 1):  # Assuming first row has headers
            if cell.value == col_name:
                return idx
        return None

    def auto_fit_columns(self, ws):
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

# Create the main window
if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelMapperApp(root)
    root.mainloop()
