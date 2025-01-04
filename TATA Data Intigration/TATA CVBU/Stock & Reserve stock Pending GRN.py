import os
import pandas as pd
from tkinter import Tk, Label, Button, filedialog, messagebox, StringVar, BooleanVar, Checkbutton, Frame
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def load_location_mapping(mapping_file):
    try:
        mapping_df = pd.read_excel(mapping_file, engine='openpyxl')
        if 'Code' not in mapping_df.columns or 'Final Location' not in mapping_df.columns:
            raise ValueError("Mapping file must contain 'Code' and 'Final Location' columns.")
        return dict(zip(mapping_df['Code'], mapping_df['Final Location']))
    except Exception as e:
        messagebox.showerror("Error", f"Error loading location mapping file: {e}")
        return {}

def process_file(file_path, location_mapping=None):
    try:
        df = pd.read_excel(file_path, engine='openpyxl')
        if location_mapping:
            df['Location'] = df['Location'].map(lambda x: location_mapping.get(x, x))
        return df
    except Exception as e:
        messagebox.showwarning("Warning", f"Error reading file {file_path}: {e}")
        return pd.DataFrame()

def generate_reports(input_folder, output_stock_folder, output_reserve_folder, pending_grn_folder, mapping_file, generate_stock, generate_reserve, generate_pending_grn):
    location_mapping = load_location_mapping(mapping_file) if mapping_file else {}

    file_paths = [os.path.join(input_folder, file) for file in os.listdir(input_folder) if file.endswith('.xlsx')]
    if not file_paths:
        messagebox.showwarning("Warning", "No Excel files found in the selected folder.")
        return

    compiled_df = pd.concat([process_file(file, location_mapping) for file in file_paths], ignore_index=True)

    if generate_stock and output_stock_folder:
        stock_columns = ['Part #', 'Qty', 'Inventory Location']
        if all(col in compiled_df.columns for col in stock_columns):
            stock_df = compiled_df[(compiled_df['Availability'] == 'On Hand') & (compiled_df['Status'] == 'Good')]
            stock_df = stock_df[stock_columns]
            stock_df = stock_df.rename(columns={
                'Part #': 'Partnumber',
                'Inventory Location': 'Location'
            })
            stock_df['Partnumber'] = stock_df['Partnumber'].astype(str)
            stock_df['Location'] = stock_df['Location'].map(lambda x: location_mapping.get(x, x) if location_mapping else x)
            save_and_format(os.path.join(output_stock_folder, "stock_upload.xlsx"), stock_df)
        else:
            messagebox.showerror("Error", "Required columns for stock upload are missing.")

    if generate_reserve and output_reserve_folder:
        reserve_df = compiled_df[compiled_df['Availability'] == 'Reserved']
        save_and_format(os.path.join(output_reserve_folder, "reserve_stock.xlsx"), reserve_df)

    if generate_pending_grn and pending_grn_folder:
        pending_grn_df = compiled_df[compiled_df['Status'] == 'In Transit']
        save_and_format(os.path.join(pending_grn_folder, "Pending_GRN.xlsx"), pending_grn_df)

def save_and_format(file_path, df):
    try:
        df.to_excel(file_path, index=False, engine='openpyxl')
        wb = load_workbook(file_path)
        ws = wb.active

        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        wb.save(file_path)
        messagebox.showinfo("Success", f"Excel file '{file_path}' has been created and formatted successfully.")
    except Exception as e:
        messagebox.showerror("Error", f"Error saving file {file_path}: {e}")

def select_input_folder():
    folder_selected = filedialog.askdirectory(title="Select Input Folder")
    if folder_selected:
        input_folder_var.set(folder_selected)

def select_stock_folder():
    folder_selected = filedialog.askdirectory(title="Select Output Folder for Stock Upload")
    if folder_selected:
        output_stock_folder_var.set(folder_selected)

def select_reserve_folder():
    folder_selected = filedialog.askdirectory(title="Select Output Folder for Reserve Stock")
    if folder_selected:
        output_reserve_folder_var.set(folder_selected)

def select_pending_grn_folder():
    folder_selected = filedialog.askdirectory(title="Select Output Folder for Pending GRN")
    if folder_selected:
        pending_grn_folder_var.set(folder_selected)

def select_mapping_file():
    file_selected = filedialog.askopenfilename(defaultextension=".xlsx",
                                              filetypes=[("Excel files", "*.xlsx")],
                                              title="Select Location Mapping File")
    if file_selected:
        mapping_file_var.set(file_selected)

def process_files():
    input_folder = input_folder_var.get()
    output_stock_folder = output_stock_folder_var.get()
    output_reserve_folder = output_reserve_folder_var.get()
    pending_grn_folder = pending_grn_folder_var.get()
    mapping_file = mapping_file_var.get()
    generate_stock = stock_var.get()
    generate_reserve = reserve_var.get()
    generate_pending_grn = pending_grn_var.get()

    if not input_folder:
        messagebox.showwarning("Warning", "Please select an input folder.")
        return

    if not (generate_stock or generate_reserve or generate_pending_grn):
        messagebox.showwarning("Warning", "Please select at least one report type to generate.")
        return

    try:
        generate_reports(input_folder, output_stock_folder, output_reserve_folder, pending_grn_folder, mapping_file, generate_stock, generate_reserve, generate_pending_grn)
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

# Create the main window
root = Tk()
root.title("Excel File Processor")
root.geometry("700x500")
root.resizable(False, False)

# Create and place widgets in frames for better layout
frame1 = Frame(root, padx=10, pady=10)
frame1.pack(pady=10, fill='x')

Label(frame1, text="Select Input Folder:", font=('Arial', 10, 'bold')).grid(row=0, column=0, sticky='w', padx=5, pady=5)
input_folder_var = StringVar()
Button(frame1, text="Browse", command=select_input_folder, relief='raised', width=20).grid(row=0, column=1, padx=5, pady=5)
Label(frame1, textvariable=input_folder_var, wraplength=400).grid(row=0, column=2, padx=5, pady=5, sticky='w')

Label(frame1, text="Select Output Folder for Stock Upload:", font=('Arial', 10, 'bold')).grid(row=1, column=0, sticky='w', padx=5, pady=5)
output_stock_folder_var = StringVar()
Button(frame1, text="Browse", command=select_stock_folder, relief='raised', width=20).grid(row=1, column=1, padx=5, pady=5)
Label(frame1, textvariable=output_stock_folder_var, wraplength=400).grid(row=1, column=2, padx=5, pady=5, sticky='w')

Label(frame1, text="Select Output Folder for Reserve Stock:", font=('Arial', 10, 'bold')).grid(row=2, column=0, sticky='w', padx=5, pady=5)
output_reserve_folder_var = StringVar()
Button(frame1, text="Browse", command=select_reserve_folder, relief='raised', width=20).grid(row=2, column=1, padx=5, pady=5)
Label(frame1, textvariable=output_reserve_folder_var, wraplength=400).grid(row=2, column=2, padx=5, pady=5, sticky='w')

Label(frame1, text="Select Output Folder for Pending GRN:", font=('Arial', 10, 'bold')).grid(row=3, column=0, sticky='w', padx=5, pady=5)
pending_grn_folder_var = StringVar()
Button(frame1, text="Browse", command=select_pending_grn_folder, relief='raised', width=20).grid(row=3, column=1, padx=5, pady=5)
Label(frame1, textvariable=pending_grn_folder_var, wraplength=400).grid(row=3, column=2, padx=5, pady=5, sticky='w')

Label(frame1, text="Select Location Mapping File (optional):", font=('Arial', 10, 'bold')).grid(row=4, column=0, sticky='w', padx=5, pady=5)
mapping_file_var = StringVar()
Button(frame1, text="Browse", command=select_mapping_file, relief='raised', width=20).grid(row=4, column=1, padx=5, pady=5)
Label(frame1, textvariable=mapping_file_var, wraplength=400).grid(row=4, column=2, padx=5, pady=5, sticky='w')

frame2 = Frame(root, padx=10, pady=10)
frame2.pack(pady=10, fill='x')

stock_var = BooleanVar()
Checkbutton(frame2, text="Generate Stock Upload Report", variable=stock_var, font=('Arial', 10)).grid(row=0, column=0, sticky='w', padx=5, pady=5)

reserve_var = BooleanVar()
Checkbutton(frame2, text="Generate Reserve Stock Report", variable=reserve_var, font=('Arial', 10)).grid(row=1, column=0, sticky='w', padx=5, pady=5)

pending_grn_var = BooleanVar()
Checkbutton(frame2, text="Generate Pending GRN Report", variable=pending_grn_var, font=('Arial', 10)).grid(row=2, column=0, sticky='w', padx=5, pady=5)

frame3 = Frame(root, padx=10, pady=10)
frame3.pack(pady=10)

Button(frame3, text="Run Process", command=process_files, relief='raised', width=20, font=('Arial', 12)).pack(padx=10, pady=10)

root.mainloop()
