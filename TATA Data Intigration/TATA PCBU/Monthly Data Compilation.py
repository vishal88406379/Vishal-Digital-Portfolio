import os
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

def compile_excel_files(folder_path, output_file):
    try:
        files = os.listdir(folder_path)
    except FileNotFoundError:
        messagebox.showerror("TATA SCS Monthly Data", "The folder path does not exist.")
        return
    except PermissionError:
        messagebox.showerror("TATA SCS Monthly Data", "Permission denied to access the folder.")
        return

    excel_files = [f for f in files if f.endswith('.xlsx')]

    if not excel_files:
        messagebox.showwarning("TATA SCS Monthly Data", "No Excel files found in the folder.")
        return

    compiled_df = pd.DataFrame()
    all_columns = set()

    for file in excel_files:
        input_file = os.path.join(folder_path, file)

        try:
            wb = load_workbook(input_file, data_only=True)
            ws = wb.active

            # Load data into a DataFrame
            data = ws.values
            columns = next(data)  # Assuming first row is header
            df = pd.DataFrame(data, columns=columns)

            # Add columns to the set of all columns
            all_columns.update(df.columns)
            compiled_df = pd.concat([compiled_df, df], ignore_index=True, sort=False)
        except Exception as e:
            messagebox.showerror("TATA SCS Monthly Data", f"Error processing file {file}: {e}")
            continue

    # Ensure all columns are present in the final DataFrame
    for col in all_columns:
        if col not in compiled_df.columns:
            compiled_df[col] = pd.NA

    # Convert date columns to datetime for accurate comparison (customize based on your date columns)
    date_columns = ['Date Column 1', 'Date Column 2']  # Replace with your actual date column names
    for col in date_columns:
        if col in compiled_df.columns:
            compiled_df[col] = pd.to_datetime(compiled_df[col], errors='coerce', format='%d-%m-%Y %H:%M')

    try:
        wb = Workbook()
        ws = wb.active

        for r in dataframe_to_rows(compiled_df, index=False, header=True):
            ws.append(r)

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        for cell in ws[1]:
            cell.font = Font(bold=False)

        wb.save(output_file)
        messagebox.showinfo("TATA SCS Monthly Data", f"File saved successfully as {output_file}")

    except Exception as e:
        messagebox.showerror("TATA SCS Monthly Data", f"Error saving file: {e}")

def browse_folder(radio_value):
    folder_path = filedialog.askdirectory()
    folder_entries[radio_value].delete(0, tk.END)
    folder_entries[radio_value].insert(0, folder_path)

def browse_output_file(radio_value):
    output_file = filedialog.asksaveasfilename(defaultextension=".xlsx")
    output_entries[radio_value].delete(0, tk.END)
    output_entries[radio_value].insert(0, output_file)

def run_compile(radio_value):
    folder_path = folder_entries[radio_value].get()
    output_file = output_entries[radio_value].get()

    # Check if the checkbox is selected
    if name_check_vars[radio_value].get():
        output_file = os.path.join(os.path.dirname(output_file), f"{keywords[radio_value]}.xlsx")
    
    compile_excel_files(folder_path, output_file)

def run_all():
    for idx, keyword in enumerate(keywords):
        folder_path = folder_entries[idx].get()
        output_file = output_entries[idx].get()

        if folder_path and output_file:
            # Determine if the file should be saved with keyword name
            if name_check_vars[idx].get():
                output_file = os.path.join(os.path.dirname(output_file), f"{keyword}.xlsx")
            
            compile_excel_files(folder_path, output_file)
        else:
            messagebox.showwarning("TATA SCS Monthly Data", f"Skipping keyword '{keyword}' due to missing folder or output path.")

    messagebox.showinfo("TATA SCS Monthly Data", "Selected files processed successfully.")

# Create the main window
root = tk.Tk()
root.title("TATA SCS Excel File Compiler")

# Create frames for better layout
frame_header = tk.Frame(root, padx=10, pady=10)
frame_header.pack(fill=tk.X)

frame_inputs = tk.Frame(root, padx=10, pady=10)
frame_inputs.pack(fill=tk.BOTH, expand=True)

frame_buttons = tk.Frame(root, padx=10, pady=10)
frame_buttons.pack(fill=tk.X)

# Header Label
header_label = tk.Label(frame_header, text="TATA SCS Monthly Data Compiler", font=("Arial", 16, "bold"))
header_label.pack()

# Keywords selection and control layout
keywords = [
    "OTC INVOICE", "PURCHASE LINE PO", "Purchase Line Items",
    "SPARE CONSUMPTION", "CLOSING STOCK", "stock transaction",
    "channel partner", "Job Line Invoice", "sap purchase order reason"
]

folder_entries = {}
output_entries = {}
name_check_vars = {}

for idx, keyword in enumerate(keywords):
    tk.Label(frame_inputs, text=f"{keyword}:", font=("Arial", 12, "bold")).grid(row=idx, column=0, padx=5, pady=5, sticky="w")

    tk.Label(frame_inputs, text="Folder:", font=("Arial", 10)).grid(row=idx, column=1, padx=5, pady=5, sticky="w")
    folder_entry = tk.Entry(frame_inputs, width=40)
    folder_entry.grid(row=idx, column=2, padx=5, pady=5, sticky="ew")
    folder_entries[idx] = folder_entry
    browse_folder_button = tk.Button(frame_inputs, text="Browse Folder", command=lambda v=idx: browse_folder(v), bg="lightblue")
    browse_folder_button.grid(row=idx, column=3, padx=5, pady=5)

    tk.Label(frame_inputs, text="Output:", font=("Arial", 10)).grid(row=idx, column=4, padx=5, pady=5, sticky="w")
    output_entry = tk.Entry(frame_inputs, width=40)
    output_entry.grid(row=idx, column=5, padx=5, pady=5, sticky="ew")
    output_entries[idx] = output_entry
    browse_output_button = tk.Button(frame_inputs, text="Browse Output", command=lambda v=idx: browse_output_file(v), bg="lightgreen")
    browse_output_button.grid(row=idx, column=6, padx=5, pady=5)

    # Checkbox for naming option
    name_check_var = tk.BooleanVar()
    name_check_vars[idx] = name_check_var
    name_checkbox = tk.Checkbutton(frame_inputs, text="Save with keyword name", variable=name_check_var)
    name_checkbox.grid(row=idx, column=7, padx=5, pady=5, sticky="w")

    run_button = tk.Button(frame_inputs, text=f"Run for {keyword}", command=lambda v=idx: run_compile(v), bg="blue", fg="white", font=('Arial', 10, 'bold'))
    run_button.grid(row=idx, column=8, padx=5, pady=5)

# Run All button
run_all_button = tk.Button(frame_buttons, text="Run All", command=run_all, bg="blue", fg="white", font=('Arial', 12, 'bold'))
run_all_button.pack()

# Checkbox for save with keyword name option
save_as_keyword_var = tk.BooleanVar()
save_as_keyword_checkbox = tk.Checkbutton(frame_buttons, text="Save files with keyword name", variable=save_as_keyword_var)
save_as_keyword_checkbox.pack()

# Configure column weights for expanding
frame_inputs.grid_columnconfigure(2, weight=1)
frame_inputs.grid_columnconfigure(5, weight=1)

# Start the GUI event loop
root.mainloop()
