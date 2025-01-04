import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle

def select_sap_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls"), ("CSV files", "*.csv")])
    sap_entry.delete(0, tk.END)
    sap_entry.insert(0, file_path)

def select_intransit_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls"), ("CSV files", "*.csv")])
    intransit_entry.delete(0, tk.END)
    intransit_entry.insert(0, file_path)

def select_output_path():
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv")])
    output_entry.delete(0, tk.END)
    output_entry.insert(0, file_path)

def get_output_file_name(base_name, extension):
    """ Generate a new file name by checking existing files in the directory. """
    counter = 1
    new_file_name = f"{base_name} {counter}.{extension}"
    while os.path.exists(new_file_name):
        counter += 1
        new_file_name = f"{base_name} {counter}.{extension}"
    return new_file_name

def auto_fit_columns(file_path):
    """ Auto-fit columns in the Excel file. """
    workbook = load_workbook(file_path)
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter  # Get the column letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)  # Add some padding
            worksheet.column_dimensions[column_letter].width = adjusted_width
    workbook.save(file_path)

def set_column_format_as_text(file_path, column_name):
    """ Set the specified column's format to text. """
    workbook = load_workbook(file_path)
    worksheet = workbook.active

    text_style = NamedStyle(name='text_style', number_format='@')  # Text format

    for row in range(2, worksheet.max_row + 1):  # Start from row 2 to skip header
        cell = worksheet[f"{column_name}{row}"]
        cell.style = text_style  # Apply text style

    workbook.save(file_path)

def run_merge():
    sap_file = sap_entry.get()
    intransit_file = intransit_entry.get()
    output_path = output_entry.get()

    if not sap_file or not output_path:
        messagebox.showerror("Error", "Please select the SAP file and an output path.")
        return

    try:
        # Read SAP Purchase Back Order
        sap_df = pd.read_excel(sap_file) if sap_file.endswith(('.xlsx', '.xls')) else pd.read_csv(sap_file)

        # Prepare SAP DataFrame
        sap_df = sap_df.rename(columns={
            'SAP Order Num': 'OrderNumber',
            'Order Date': 'OrderDate',
            'Part No': 'PartNumber',
            'Pending Qty.': 'POQty'
        })[['Location', 'OrderNumber', 'OrderDate', 'PartNumber', 'POQty']]

        # Handle Intransit file
        if intransit_file:
            intransit_df = pd.read_excel(intransit_file) if intransit_file.endswith(('.xlsx', '.xls')) else pd.read_csv(intransit_file)
            intransit_df = intransit_df.rename(columns={
                'Supplied qty': 'POQty'
            })[['Location', 'OrderNumber', 'OrderDate', 'PartNumber', 'POQty']]
        else:
            intransit_df = pd.DataFrame(columns=['Location', 'OrderNumber', 'OrderDate', 'PartNumber', 'POQty'])

        # Concatenate both DataFrames
        merged_df = pd.concat([sap_df, intransit_df], ignore_index=True)

        # Add blank columns
        merged_df['OEMInvoiceNo'] = ''
        merged_df['OEMInvoiceDate'] = ''
        merged_df['OEMInvoiceQty'] = ''

        # Get the output file name with incremental number
        output_base_name = "Sm auto All Oem"
        output_extension = "xlsx"
        output_dir = os.path.dirname(output_path) if output_path else '.'
        output_file = get_output_file_name(os.path.join(output_dir, output_base_name), output_extension)

        # Save merged output to a single file with sheet name "Sheet1"
        merged_df.to_excel(output_file, index=False, sheet_name='Sheet1')
        auto_fit_columns(output_file)  # Auto-fit columns in the merged file
        set_column_format_as_text(output_file, 'D')  # Set 'PartNumber' column (D) format to text
        messagebox.showinfo("Success", f"Merged file created successfully: {output_file}")

        # Define retail and tass locations
        retail_locations = [
            'Varanasi Retail',
            'Lucknow STU Retail',
            'Lucknow Retail',
            'Gorakhpur Retail',
            'Allahabad Retail',
            'Chopan Retail',
            'Faizabad Retail',
            'Fatehpur Retail'
        ]

        tass_locations = [
            'Varanasi TASS',
            'LucknowES Tass'
        ]

        # Create Retail file
        retail_df = merged_df[merged_df['Location'].isin(retail_locations)]
        retail_file = os.path.join(output_dir, 'OemInvoice Retail.xlsx')
        retail_df.to_excel(retail_file, index=False, sheet_name='Sheet1')
        auto_fit_columns(retail_file)  # Auto-fit columns in the retail file
        set_column_format_as_text(retail_file, 'D')  # Set 'PartNumber' column (D) format to text
        print(f"Created file for retail locations: {retail_file}")

        # Create Tass file
        tass_df = merged_df[merged_df['Location'].isin(tass_locations)]
        tass_file = os.path.join(output_dir, 'OemInvoice Tass.xlsx')
        tass_df.to_excel(tass_file, index=False, sheet_name='Sheet1')
        auto_fit_columns(tass_file)  # Auto-fit columns in the tass file
        set_column_format_as_text(tass_file, 'D')  # Set 'PartNumber' column (D) format to text
        print(f"Created file for tass locations: {tass_file}")

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

# GUI setup
root = tk.Tk()
root.title("Data Merge Tool")

tk.Label(root, text="SAP Purchase Back Order File:").grid(row=0, column=0, padx=10, pady=10)
sap_entry = tk.Entry(root, width=50)
sap_entry.grid(row=0, column=1, padx=10, pady=10)
tk.Button(root, text="Browse", command=select_sap_file).grid(row=0, column=2, padx=10, pady=10)

tk.Label(root, text="Intransit File (Optional):").grid(row=1, column=0, padx=10, pady=10)
intransit_entry = tk.Entry(root, width=50)
intransit_entry.grid(row=1, column=1, padx=10, pady=10)
tk.Button(root, text="Browse", command=select_intransit_file).grid(row=1, column=2, padx=10, pady=10)

tk.Label(root, text="Output Path:").grid(row=2, column=0, padx=10, pady=10)
output_entry = tk.Entry(root, width=50)
output_entry.grid(row=2, column=1, padx=10, pady=10)
tk.Button(root, text="Browse", command=select_output_path).grid(row=2, column=2, padx=10, pady=10)

tk.Button(root, text="Run", command=run_merge).grid(row=3, columnspan=3, padx=10, pady=20)

root.mainloop()
