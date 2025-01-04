import os
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

def compile_excel_files(folder_path, output_file):
    try:
        files = os.listdir(folder_path)
        excel_files = [f for f in files if f.endswith('.xlsx')]
        if not excel_files:
            raise FileNotFoundError("No Excel files found in the folder.")
        
        # Use a list to hold the DataFrames and their headers
        dfs = []
        headers = set()
        
        for file in excel_files:
            input_file = os.path.join(folder_path, file)
            try:
                wb = load_workbook(input_file, data_only=False)  # Load with original formatting
                ws = wb.active
                data = ws.values
                columns = next(data)  # First row is header
                df = pd.DataFrame(data, columns=columns)
                
                # Collect headers
                headers.update(df.columns)
                dfs.append(df)
            except Exception as e:
                print(f"Error processing file {file}: {e}")
                continue

        # Combine all DataFrames into one
        compiled_df = pd.concat(dfs, ignore_index=True, sort=False)
        
        # Ensure all columns are present
        for header in headers:
            if header not in compiled_df.columns:
                compiled_df[header] = pd.NA

        # Save the compiled DataFrame to an Excel file
        wb = Workbook()
        ws = wb.active
        
        # Append DataFrame to the worksheet
        for r in dataframe_to_rows(compiled_df, index=False, header=True):
            ws.append(r)

        # Adjust column widths based on content
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Remove bold formatting from header row
        for cell in ws[1]:
            cell.font = Font(bold=False)

        wb.save(output_file)
        messagebox.showinfo("Success", "Excel files compiled successfully!")
    
    except FileNotFoundError as e:
        messagebox.showerror("Error", f"{e}")
    except PermissionError as e:
        messagebox.showerror("Error", f"Permission error: {e}")
    except Exception as e:
        messagebox.showerror("Error", f"An unexpected error occurred: {e}")

def browse_folder():
    folder_path = filedialog.askdirectory()
    folder_entry.delete(0, tk.END)
    folder_entry.insert(0, folder_path)

def browse_output_file():
    output_file = filedialog.asksaveasfilename(defaultextension=".xlsx")
    output_entry.delete(0, tk.END)
    output_entry.insert(0, output_file)

def run_compile():
    folder_path = folder_entry.get()
    output_file = output_entry.get()
    try:
        compile_excel_files(folder_path, output_file)
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

root = tk.Tk()
root.title("Excel File Compiler")

folder_label = tk.Label(root, text="Select Folder:")
folder_label.pack()

folder_entry = tk.Entry(root, width=50)
folder_entry.pack()

browse_folder_button = tk.Button(root, text="Browse", command=browse_folder)
browse_folder_button.pack()

output_label = tk.Label(root, text="Select Output File:")
output_label.pack()

output_entry = tk.Entry(root, width=50)
output_entry.pack()

browse_output_button = tk.Button(root, text="Browse", command=browse_output_file)
browse_output_button.pack()

run_button = tk.Button(root, text="Run", command=run_compile)
run_button.pack()

root.mainloop()
