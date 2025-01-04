import os
import pandas as pd
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle
import logging
import openpyxl  # Ensure openpyxl is imported

# Set up logging
logging.basicConfig(filename='processing.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def read_file(file_path):
    """Read file based on its extension."""
    file_path_str = str(file_path)  # Ensure file path is a string
    try:
        if file_path_str.endswith('.xlsx'):
            return pd.read_excel(file_path)
        elif file_path_str.endswith('.csv'):
            return pd.read_csv(file_path)
        elif file_path_str.endswith('.xls'):
            return pd.read_excel(file_path, engine='xlrd')
        elif file_path_str.endswith('.json'):
            return pd.read_json(file_path)
        else:
            raise ValueError("Unsupported file format")
    except Exception as e:
        logging.error(f"Failed to read file {file_path}: {e}")
        raise

def validate_columns(df, required_columns):
    """Validate if the required columns are in the DataFrame."""
    return all(column in df.columns for column in required_columns)

def load_location_mapping(mapping_file):
    """Load location mapping from a file."""
    location_df = read_file(mapping_file)
    
    # Check if the required columns exist in the dataframe
    if 'Code' not in location_df.columns or 'Final Location' not in location_df.columns:
        raise ValueError("Mapping file must contain 'Code' and 'Final Location' columns.")
    
    return dict(zip(location_df['Code'], location_df['Final Location']))

def process_sap_purchase_backorders(file_paths, max_days_pending, location_mapping):
    """Process SAP purchase backorder files."""
    combined_df = pd.DataFrame()
    required_columns = ['Division', 'Order Number', 'Order Date', 'Part No', 'Pending Qty.']
    
    for file_path in file_paths:
        try:
            df = read_file(file_path)
            if not validate_columns(df, required_columns):
                raise ValueError(f"File {file_path} does not contain required columns.")
            
            df.rename(columns={
                'Division': 'Location',
                'Order Number': 'OrderNumber',
                'Order Date': 'OrderDate',
                'Part No': 'PartNumber',
                'Pending Qty.': 'POQty'
            }, inplace=True)
            
            df['PartNumber'] = df['PartNumber'].astype(str)
            df['Location'] = df['Location'].map(location_mapping).combine_first(df['Location'])
            df['Days Pending'] = (pd.Timestamp.today() - pd.to_datetime(df['OrderDate'])).dt.days
            df_filtered = df[df['Days Pending'] <= max_days_pending]
            combined_df = pd.concat([combined_df, df_filtered[['Location', 'OrderNumber', 'OrderDate', 'PartNumber', 'POQty']]], ignore_index=True)
        except Exception as e:
            logging.error(f"Error processing SAP file {file_path}: {e}")
    
    return combined_df

def process_intransit_files(folder_path, max_invoice_days, location_mapping):
    """Process intransit files."""
    combined_df = pd.DataFrame()
    
    for file_path in Path(folder_path).glob("*"):
        try:
            df = read_file(file_path)
            df.rename(columns={
                'Division Name': 'Location',
                'Order #': 'OrderNumber',
                'Part #': 'PartNumber',
                'Recd Qty': 'POQty',
                'Invoice_Date': 'OrderDate'
            }, inplace=True)
            
            df['PartNumber'] = df['PartNumber'].astype(str)
            df['Days Since Invoice'] = (pd.Timestamp.today() - pd.to_datetime(df['OrderDate'])).dt.days
            df_filtered = df[
                (df['Days Since Invoice'] > 0) & 
                (df['Days Since Invoice'] < max_invoice_days) & 
                (df['Status'] == 'In Transit')
            ]
            
            df_filtered = df_filtered.dropna(how='all', axis=1)
            df_filtered['Location'] = df_filtered['Location'].map(location_mapping).combine_first(df_filtered['Location'])
            combined_df = pd.concat([combined_df, df_filtered], ignore_index=True)
        except Exception as e:
            logging.error(f"Error processing Intransit file {file_path}: {e}")
    
    combined_df = combined_df.dropna(how='all', axis=1)
    
    return combined_df[['Location', 'OrderNumber', 'OrderDate', 'PartNumber', 'POQty']]

def save_output(final_output, output_folder_path, file_format='xlsx'):
    """Save the output to a specified file format."""
    try:
        if file_format == 'xlsx':
            output_file_path = os.path.join(output_folder_path, "OEMInvoice.xlsx")
            with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
                final_output.to_excel(writer, index=False, sheet_name='Sheet1')
                workbook = writer.book
                worksheet = writer.sheets['Sheet1']
                
                # Apply borders around all cells
                thin_border = Border(left=Side(border_style='thin'), 
                                     right=Side(border_style='thin'),
                                     top=Side(border_style='thin'),
                                     bottom=Side(border_style='thin'))
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.border = thin_border

                # Adjust column width to fit data
                for col in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(col[0].column)
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                # Ensure no header formatting
                for cell in worksheet[1]:
                    cell.font = Font(bold=False)  # Ensure headers are not bold
                    cell.fill = openpyxl.styles.PatternFill(fill_type=None)  # No fill color

                # Ensure PartNumber is treated as text
                part_number_style = NamedStyle(name="part_number_style", number_format="@")
                for cell in worksheet['D']:  # Assuming PartNumber is in column 'D'
                    cell.style = part_number_style

                # Save the workbook
                workbook.save(output_file_path)
        elif file_format == 'csv':
            output_file_path = os.path.join(output_folder_path, "OEMInvoice.csv")
            final_output.to_csv(output_file_path, index=False)
        else:
            raise ValueError("Unsupported file format")
        
        logging.info(f"Combined and filtered data saved to {output_folder_path}/OEMInvoice.{file_format}")
        messagebox.showinfo("Success", f"Combined and filtered data saved to {output_folder_path}/OEMInvoice.{file_format}")
    except Exception as e:
        logging.error(f"Error saving output file: {e}")
        messagebox.showerror("Error", f"Failed to save output file: {e}")

def combine_and_save_output(sap_files, intransit_folder_path, output_folder_path, mapping_file, max_days_pending=35, max_invoice_days=90, file_format='xlsx'):
    """Combine and save the final output."""
    try:
        location_mapping = load_location_mapping(mapping_file)
        
        # Process SAP files if provided
        if sap_files:
            sap_output = process_sap_purchase_backorders(sap_files, max_days_pending, location_mapping)
        else:
            sap_output = pd.DataFrame(columns=['Location', 'OrderNumber', 'OrderDate', 'PartNumber', 'POQty'])
        
        # Process Intransit files if folder path is provided
        if intransit_folder_path:
            intransit_output = process_intransit_files(intransit_folder_path, max_invoice_days, location_mapping)
        else:
            intransit_output = pd.DataFrame(columns=['Location', 'OrderNumber', 'OrderDate', 'PartNumber', 'POQty'])
        
        # Combine the outputs
        final_output = pd.concat([sap_output, intransit_output], ignore_index=True)
        final_output['OEMInvoiceNo'] = ''
        final_output['OEMInvoiceDate'] = ''
        final_output['OEMInvoiceQty'] = ''

        # Save the output
        save_output(final_output, output_folder_path, file_format)
    except ValueError as ve:
        messagebox.showerror("Error", str(ve))
    except Exception as e:
        logging.error(f"Error occurred: {e}")
        messagebox.showerror("Error", str(e))

def run_gui():
    """Set up and run the GUI."""
    def select_sap_files():
        sap_paths = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv"), ("JSON files", "*.json")])
        sap_files_var.set(';'.join(sap_paths))

    def select_intransit_folder():
        intransit_path = filedialog.askdirectory()
        intransit_folder_var.set(intransit_path)

    def select_output_folder():
        output_path = filedialog.askdirectory()
        output_folder_var.set(output_path)

    def select_mapping_file():
        mapping_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv"), ("JSON files", "*.json")])
        mapping_file_var.set(mapping_path)

    def run_processing():
        sap_files = sap_files_var.get().split(';')
        intransit_path = intransit_folder_var.get()
        output_path = output_folder_var.get()
        mapping_file = mapping_file_var.get()

        try:
            combine_and_save_output(sap_files, intransit_path, output_path, mapping_file)
        except Exception as e:
            messagebox.showerror("Error", str(e))

    # Set up GUI
    root = tk.Tk()
    root.title("Data Processing")

    tk.Label(root, text="Select SAP Purchase Backorder Files:").pack()
    sap_files_var = tk.StringVar()
    tk.Entry(root, textvariable=sap_files_var, width=100).pack()
    tk.Button(root, text="Browse", command=select_sap_files).pack()

    tk.Label(root, text="Select Intransit Folder:").pack()
    intransit_folder_var = tk.StringVar()
    tk.Entry(root, textvariable=intransit_folder_var, width=100).pack()
    tk.Button(root, text="Browse", command=select_intransit_folder).pack()

    tk.Label(root, text="Select Output Folder:").pack()
    output_folder_var = tk.StringVar()
    tk.Entry(root, textvariable=output_folder_var, width=100).pack()
    tk.Button(root, text="Browse", command=select_output_folder).pack()

    tk.Label(root, text="Select Location Mapping File:").pack()
    mapping_file_var = tk.StringVar()
    tk.Entry(root, textvariable=mapping_file_var, width=100).pack()
    tk.Button(root, text="Browse", command=select_mapping_file).pack()

    tk.Button(root, text="Run Processing", command=run_processing).pack()

    root.mainloop()

if __name__ == "__main__":
    run_gui()
