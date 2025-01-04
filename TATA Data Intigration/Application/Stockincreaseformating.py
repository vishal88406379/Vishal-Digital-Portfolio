import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import logging
import pandas as pd

# Set up logging
logging.basicConfig(filename='excel_processor.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def format_header(sheet):
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    bold_font = Font(bold=True)

    for cell in sheet[1]:  # Assuming the first row is the header
        cell.font = bold_font
        cell.fill = yellow_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

def apply_borders(sheet):
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in sheet.iter_rows():
        for cell in row:
            cell.border = thin_border

def auto_fit_columns(sheet):
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Get the column letter (A, B, C, ...)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception as e:
                logging.error(f"Error calculating width for cell {cell.coordinate}: {e}")
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

def process_excel_file(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            for row in sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            format_header(sheet)
            apply_borders(sheet)
            auto_fit_columns(sheet)
        
        workbook.save(file_path)
        logging.info(f"Successfully processed: {file_path}")
    
    except PermissionError:
        logging.error(f"Permission denied: {file_path}. Ensure the file is not open and has write permissions.")
        messagebox.showwarning("Warning", f"File is open or locked: {file_path}. Please close it before processing.")
    except Exception as e:
        logging.error(f"An error occurred while processing {file_path}: {e}")
        messagebox.showerror("Error", f"An error occurred while processing {file_path}. Check logs for details.")

def convert_and_process_file(file_path):
    try:
        # Convert .xls or .csv to .xlsx using pandas
        if file_path.endswith('.xls') or file_path.endswith('.csv'):
            # Read the file into a DataFrame
            df = pd.read_excel(file_path, engine='xlrd') if file_path.endswith('.xls') else pd.read_csv(file_path)
            
            # Save the DataFrame as an .xlsx file
            temp_xlsx = file_path.rsplit('.', 1)[0] + '.xlsx'
            df.to_excel(temp_xlsx, index=False, engine='openpyxl')
            
            logging.info(f"Converted {file_path} to {temp_xlsx}")
            
            # Process the newly created .xlsx file
            process_excel_file(temp_xlsx)
            
            # Optionally, remove the temporary .xlsx file
            os.remove(temp_xlsx)
        else:
            process_excel_file(file_path)
    except Exception as e:
        logging.error(f"An error occurred during file conversion or processing {file_path}: {e}")
        messagebox.showerror("Error", f"An error occurred with {file_path}. Check logs for details.")

def process_folder(folder_path):
    files = [f for f in os.listdir(folder_path) if f.endswith(('.xlsx', '.xls', '.csv'))]
    
    for index, filename in enumerate(files, start=1):
        file_path = os.path.join(folder_path, filename)
        convert_and_process_file(file_path)
        progress_var.set((index / len(files)) * 100)
        root.update_idletasks()  # Update progress bar

    messagebox.showinfo("Success", "All files have been processed successfully!")
    logging.info("All files processed successfully.")

def browse_folder():
    folder_path = filedialog.askdirectory()
    if folder_path:
        folder_entry.delete(0, tk.END)
        folder_entry.insert(0, folder_path)

def start_processing():
    folder_path = folder_entry.get()
    if folder_path:
        process_folder(folder_path)
    else:
        messagebox.showwarning("Warning", "Please select a folder first.")

# GUI setup
root = tk.Tk()
root.title("Excel Files Processor")

# Folder selection
tk.Label(root, text="Select Folder:").grid(row=0, column=0, padx=10, pady=10)
folder_entry = tk.Entry(root, width=50)
folder_entry.grid(row=0, column=1, padx=10, pady=10)
browse_button = tk.Button(root, text="Browse", command=browse_folder)
browse_button.grid(row=0, column=2, padx=10, pady=10)

# Progress bar
progress_var = tk.DoubleVar()
progress_bar = ttk.Progressbar(root, orient='horizontal', length=400, mode='determinate', variable=progress_var)
progress_bar.grid(row=1, column=0, columnspan=3, padx=10, pady=10)

# Run button
run_button = tk.Button(root, text="Run", command=start_processing, bg="green", fg="white")
run_button.grid(row=2, column=0, columnspan=3, pady=20)

root.mainloop()
